from fastapi import FastAPI, File, UploadFile, BackgroundTasks
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from pathlib import Path
import openpyxl
from collections import defaultdict
from tempfile import NamedTemporaryFile
import os

from starlette.requests import Request

app = FastAPI()

# Настройка шаблонов
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

# Папка для временных файлов
TEMP_DIR = Path("temp")
TEMP_DIR.mkdir(exist_ok=True)

@app.get("/", response_class=HTMLResponse)
async def upload_form(request: Request):
    """Главная страница с формой для загрузки файла"""
    return templates.TemplateResponse("index.html", {"request": request})

@app.post("/process/")
async def process_excel(file: UploadFile = File(...), background_tasks: BackgroundTasks = BackgroundTasks()):
    """Обрабатывает загруженный Excel, суммирует длины кабелей на каждом листе и добавляет итог"""
    try:
        # Читаем загруженный файл во временное хранилище
        with NamedTemporaryFile(delete=False, suffix=".xlsx", dir=TEMP_DIR) as temp_file:
            temp_path = Path(temp_file.name)
            temp_file.write(await file.read())

        # Загружаем книгу дважды: с вычисленными значениями и с формулами
        wb_values = openpyxl.load_workbook(temp_path, data_only=True)  # Читаем вычисленные значения
        wb_formulas = openpyxl.load_workbook(temp_path)  # Читаем оригинальные формулы

        found_data = False  # Флаг, нашли ли данные

        # Перебираем все листы в книге
        for ws_values, ws_formulas in zip(wb_values.worksheets, wb_formulas.worksheets):
            cable_sums = defaultdict(float)
            headers = {}

            # Ищем строку с заголовками
            data_start_row = None
            for row in ws_values.iter_rows(min_row=1, max_row=40):
                for cell in row:
                    if isinstance(cell.value, (int, float)):  # Если ячейка содержит число
                        data_start_row = row[0].row
                        break
                if data_start_row:
                    break

            if not data_start_row:
                continue  # Если не нашли данных, пропускаем лист

            # Ищем заголовки
            for row in ws_values.iter_rows(min_row=1, max_row=data_start_row - 1):
                for cell in row:
                    if cell.value:
                        key = str(cell.value).strip()
                        col = cell.column
                        if "Тип" in key and "Тип" not in headers:
                            headers["Тип"] = col
                        if ("Число" in key and "сечение" in key) and "Число и сечение жил" not in headers:
                            headers["Число и сечение жил"] = col
                        if "Длина" in key and "Длина" not in headers:
                            headers["Длина"] = col

            if len(headers) < 3:
                continue  # Если заголовков недостаточно, пропускаем лист

            # Начинаем обработку данных
            for row in ws_values.iter_rows(min_row=data_start_row, values_only=False):
                cable_type = row[headers["Тип"] - 1].value if headers["Тип"] - 1 < len(row) else None
                spec = row[headers["Число и сечение жил"] - 1].value if headers["Число и сечение жил"] - 1 < len(row) else None
                length_cell = row[headers["Длина"] - 1] if headers["Длина"] - 1 < len(row) else None

                length = length_cell.value if length_cell and isinstance(length_cell.value, (int, float)) else None
                if length is not None:
                    cable_sums[(cable_type, spec)] += length

            # Добавляем итог в конец листа
            ws_formulas.append([""])  # Пустая строка
            ws_formulas.append(["Итог"])
            ws_formulas.append(["Тип", "Сечение", "Длина"])
            for (cable_type, spec), total_length in cable_sums.items():
                ws_formulas.append([cable_type, spec if spec else "Не указано", total_length])

            found_data = True

        if not found_data:
            raise ValueError("В файле не найдено нужных данных (нет нужных столбцов)")

        # Сохраняем обработанный файл
        output_path = TEMP_DIR / f"updated_{file.filename}"
        wb_formulas.save(output_path)
        wb_formulas.close()

        # Добавляем задачу на удаление файлов после отправки
        background_tasks.add_task(remove_file, temp_path)
        background_tasks.add_task(remove_file, output_path)

        return FileResponse(output_path, filename=f"Обновлённый_{file.filename}")

    except Exception as e:
        return {"error": f"Ошибка обработки файла: {str(e)}"}

async def remove_file(file_path: Path):
    """Удаление файла после отправки"""
    try:
        # Логируем перед удалением
        print(f"Попытка удалить файл {file_path}.")

        if file_path.exists():
            os.remove(file_path)
            print(f"Файл {file_path} удален.")
        else:
            print(f"Файл {file_path} не найден для удаления.")
    except Exception as e:
        print(f"Ошибка при удалении файла {file_path}: {e}")
