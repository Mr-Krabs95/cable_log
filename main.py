from fastapi import FastAPI, File, UploadFile, BackgroundTasks
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from pathlib import Path
import openpyxl
from collections import defaultdict # collections расширяет стандартные структура списков, словарей, кортежей; defaultdict - класс словаря
from tempfile import NamedTemporaryFile
import os
from starlette.requests import Request

app = FastAPI()

# Настройка шаблонов для рендеринга HTML-страниц
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

# Папка для временных файлов
TEMP_DIR = Path("temp")
TEMP_DIR.mkdir(exist_ok=True)

@app.get("/", response_class=HTMLResponse)
async def upload_form(request: Request):
    """Главная страница с формой для загрузки файла"""
    return templates.TemplateResponse("index.html", {"request": request})

@app.post("/process/")
async def process_excel(file: UploadFile = File(...), background_tasks: BackgroundTasks = BackgroundTasks()):
    """Обрабатывает загруженный Excel, суммирует длины кабелей и сохраняет итог"""
    try:
        # Записываем загруженный файл во временное хранилище
        with NamedTemporaryFile(delete=False, suffix=".xlsx", dir=TEMP_DIR) as temp_file:
            temp_path = Path(temp_file.name)
            temp_file.write(await file.read())

        # Загружаем книгу с вычисленными значениями и с формулами
        wb_values = openpyxl.load_workbook(temp_path, data_only=True)  # Читаем вычисленные значения
        wb_formulas = openpyxl.load_workbook(temp_path)  # Читаем оригинальные формулы

        found_data = False  # Флаг наличия данных

        for ws_values, ws_formulas in zip(wb_values.worksheets, wb_formulas.worksheets):
            cable_sums = defaultdict(lambda: defaultdict(float))  # Хранение суммы значений по кабелям
            headers = {}  # Заголовки таблицы
            numeric_columns = {}  # Числовые столбцы
            column_names = {}  # Названия столбцов

            # Поиск строки с заголовками
            data_start_row = None
            for row in ws_values.iter_rows(min_row=1, max_row=40):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        data_start_row = row[0].row
                        break
                if data_start_row:
                    break

            if not data_start_row:
                continue  # Пропускаем лист, если не нашли данных

            # Поиск заголовков и их индексов
            for row in ws_values.iter_rows(min_row=1, max_row=data_start_row - 1):
                for cell in row:
                    if cell.value:
                        key = str(cell.value).strip()
                        col = cell.column
                        column_names[col] = key  # Сохраняем название столбца
                        if "Тип" in key and "Тип" not in headers:
                            headers["Тип"] = col
                        if "Число" in key and "сечение" in key and "Число и сечение жил" not in headers:
                            headers["Число и сечение жил"] = col

            if len(headers) < 2:
                continue  # Пропускаем лист, если не нашли нужные заголовки

            # Определяем числовые столбцы (после сечения)
            for col in range(headers["Число и сечение жил"] + 1, ws_values.max_column + 1):
                # Поиск по всем строкам в столбце для получения значения
                column_name = None
                for row in ws_values.iter_rows(min_row=1, max_row=data_start_row - 1):
                    cell = row[col - 1]
                    if cell.value:
                        column_name = str(cell.value).strip()  # Найдено значение, считаем его названием столбца
                        break  # Останавливаем поиск, если нашли первое ненулевое значение
                
                # Если значение не найдено, назначаем дефолтное имя столбца
                if not column_name:
                    column_name = f"Столбец {col}"

                column_names[col] = column_name  # Сохраняем название столбца
                numeric_columns[col] = 0  # Инициализируем сумму для этого столбца

            # Суммируем данные
            for row in ws_values.iter_rows(min_row=data_start_row, values_only=False):
                cable_type = row[headers["Тип"] - 1].value
                spec = row[headers["Число и сечение жил"] - 1].value
                cable_key = (cable_type, spec)
                
                for col, _ in numeric_columns.items():
                    value = row[col - 1].value
                    if isinstance(value, (int, float)):
                        cable_sums[cable_key][col] += value

            # Добавляем итог в конец листа
            ws_formulas.append([""])  # Пустая строка
            ws_formulas.append(["Итог"])
            header_row = ["Тип", "Сечение"] + [column_names[col] for col in numeric_columns if any(cable_sums[key][col] != 0 for key in cable_sums)]
            ws_formulas.append(header_row)
            
            for (cable_type, spec), sums in cable_sums.items():
                row_data = [cable_type, spec if spec else "Не указано"] + [sums[col] for col in numeric_columns if any(sums[col] != 0 for sums in cable_sums.values())]
                ws_formulas.append(row_data)

            found_data = True

        if not found_data:
            raise ValueError("В файле не найдено нужных данных (нет нужных столбцов)")

        # Сохраняем обработанный файл
        output_path = TEMP_DIR / f"updated_{file.filename}"
        wb_formulas.save(output_path)
        wb_formulas.close()

        # Добавляем задачу на удаление файлов после отправки
        background_tasks.add_task(remove_file, temp_path)
        background_tasks.add_task(remove_file, output_path)

        return FileResponse(output_path, filename=f"Обновлённый_{file.filename}")

    except Exception as e:
        return {"error": f"Ошибка обработки файла: {str(e)}"}

async def remove_file(file_path: Path):
    """Удаление файла после отправки"""
    try:
        print(f"Попытка удалить файл {file_path}.")
        if file_path.exists():
            os.remove(file_path)
            print(f"Файл {file_path} удален.")
        else:
            print(f"Файл {file_path} не найден для удаления.")
    except Exception as e:
        print(f"Ошибка при удалении файла {file_path}: {e}")